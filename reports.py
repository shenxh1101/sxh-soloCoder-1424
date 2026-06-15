import os
from services import PartService, WorkOrderService, PartRepository
from datetime import datetime

EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')


def ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)


def export_parts_template(filename=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return False, "请先安装 openpyxl: pip install openpyxl"

    ensure_export_dir()
    if not filename:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"配件导入模板_{ts}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "配件数据"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["配件名称*", "配件编号*", "单价(元)*", "初始库存", "最小库存量"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    samples = [
        ("示例机油", "EX001", 88.50, 20, 5),
        ("示例滤清器", "EX002", 35.00, 15, 5),
    ]
    for i, s in enumerate(samples, 2):
        for col, val in enumerate(s, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    ws_note = wb.create_sheet("填写说明")
    notes = [
        "配件批量导入说明：",
        "1. 在【配件数据】sheet中填写数据，带*为必填项",
        "2. 配件编号不可重复，如已存在将标记为失败",
        "3. 单价必须为大于等于0的数字，库存和最小库存必须为非负整数",
        "4. 示例行可删除或保留，系统会自动识别编号为EX开头的为示例并跳过",
        "5. 初始库存留空默认0，最小库存量留空默认5",
    ]
    for i, note in enumerate(notes, 1):
        ws_note.cell(row=i, column=1, value=note)
    ws_note.column_dimensions['A'].width = 80

    wb.save(filepath)
    return True, filepath


def export_parts_list(filename=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return False, "请先安装 openpyxl: pip install openpyxl"

    ensure_export_dir()
    parts = PartService.list_parts()

    if not filename:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"配件库存清单_{ts}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "配件清单"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["ID", "配件名称", "配件编号", "单价(元)", "当前库存", "最小库存", "状态", "创建时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row = 2
    for p in parts:
        status = "库存不足" if p['stock'] < p['min_stock'] else "正常"
        ws.cell(row=row, column=1, value=p['id']).border = thin_border
        ws.cell(row=row, column=2, value=p['name']).border = thin_border
        ws.cell(row=row, column=3, value=p['code']).border = thin_border
        ws.cell(row=row, column=4, value=round(p['price'], 2)).border = thin_border
        ws.cell(row=row, column=5, value=p['stock']).border = thin_border
        ws.cell(row=row, column=6, value=p['min_stock']).border = thin_border
        ws.cell(row=row, column=7, value=status).border = thin_border
        ws.cell(row=row, column=8, value=p['created_at']).border = thin_border
        row += 1

    widths = [8, 22, 14, 12, 12, 12, 12, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(filepath)
    return True, filepath


def import_parts_from_excel(filepath):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False, "请先安装 openpyxl: pip install openpyxl", [], []

    if not os.path.exists(filepath):
        return False, f"文件不存在: {filepath}", [], []

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        return False, f"无法打开Excel文件: {str(e)}", [], []

    if "配件数据" not in wb.sheetnames:
        return False, "Excel中缺少【配件数据】工作表", [], []

    ws = wb["配件数据"]

    success_list = []
    failure_list = []

    existing_codes = set()
    all_parts = PartService.list_parts()
    for p in all_parts:
        existing_codes.add(p['code'].upper())

    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1

        if row is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue

        name = str(row[0]).strip() if row[0] is not None else ""
        code = str(row[1]).strip() if row[1] is not None else ""

        if code.upper().startswith("EX"):
            continue

        errors = []
        if not name:
            errors.append("配件名称为空")
        if not code:
            errors.append("配件编号为空")
        if code and code.upper() in existing_codes:
            errors.append(f"配件编号已存在")

        price = 0.0
        if len(row) > 2 and row[2] is not None:
            try:
                price = float(row[2])
                if price < 0:
                    errors.append("单价不能为负数")
            except (ValueError, TypeError):
                errors.append(f"单价格式错误: {row[2]}")
        else:
            errors.append("单价为空")

        stock = 0
        if len(row) > 3 and row[3] is not None and row[3] != "":
            try:
                stock = int(row[3])
                if stock < 0:
                    errors.append("初始库存不能为负数")
            except (ValueError, TypeError):
                errors.append(f"初始库存格式错误: {row[3]}")

        min_stock = 5
        if len(row) > 4 and row[4] is not None and row[4] != "":
            try:
                min_stock = int(row[4])
                if min_stock < 0:
                    errors.append("最小库存不能为负数")
            except (ValueError, TypeError):
                errors.append(f"最小库存格式错误: {row[4]}")

        if errors:
            failure_list.append({
                'row': row_num,
                'code': code,
                'name': name,
                'errors': errors
            })
            continue

        pid, err = PartService.add_part(name, code, price, stock, min_stock)
        if err:
            failure_list.append({
                'row': row_num,
                'code': code,
                'name': name,
                'errors': [err]
            })
            continue

        existing_codes.add(code.upper())
        success_list.append({
            'row': row_num,
            'id': pid,
            'code': code,
            'name': name,
            'price': price,
            'stock': stock,
            'min_stock': min_stock
        })

    return True, None, success_list, failure_list


def export_parts_transactions(start_date=None, end_date=None, filename=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return False, "请先安装 openpyxl: pip install openpyxl"

    ensure_export_dir()

    transactions = PartService.list_transactions(start_date, end_date)

    if not filename:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"配件出入库明细_{ts}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "出入库明细"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["日期时间", "配件编号", "配件名称", "类型", "数量", "关联工单", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    row = 2
    for txn in transactions:
        txn_type = "入库" if txn['type'] == 'IN' else "出库"
        ws.cell(row=row, column=1, value=txn['created_at']).border = thin_border
        ws.cell(row=row, column=2, value=txn.get('part_code', '')).border = thin_border
        ws.cell(row=row, column=3, value=txn.get('part_name', '')).border = thin_border
        ws.cell(row=row, column=4, value=txn_type).border = thin_border
        ws.cell(row=row, column=5, value=txn['quantity']).border = thin_border
        ws.cell(row=row, column=6, value=txn.get('order_id', '') or '').border = thin_border
        ws.cell(row=row, column=7, value=txn.get('note', '') or '').border = thin_border
        row += 1

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18

    wb.save(filepath)
    return True, filepath


def export_revenue_report(start_date, end_date, filename=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return False, "请先安装 openpyxl: pip install openpyxl"

    ensure_export_dir()

    stats = WorkOrderService.get_revenue_statistics(start_date, end_date)

    if not filename:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"维修收入统计_{start_date}_{end_date}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "统计汇总"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=11)
    value_font = Font(size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws_summary.merge_cells('A1:D1')
    ws_summary.cell(row=1, column=1, value="维修收入统计报表").font = title_font
    ws_summary.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws_summary.cell(row=3, column=1, value="统计期间:").font = label_font
    ws_summary.cell(row=3, column=2, value=f"{start_date} 至 {end_date}").font = value_font

    summary_data = [
        ("工单总数", stats['order_count']),
        ("总收入(元)", f"{stats['total_revenue']:.2f}"),
        ("  工时费(元)", f"{stats['total_labor']:.2f}"),
        ("  配件费(元)", f"{stats['total_parts_cost']:.2f}"),
        ("平均客单价(元)", f"{stats['avg_order_value']:.2f}"),
    ]

    row = 5
    for label, value in summary_data:
        ws_summary.cell(row=row, column=1, value=label).font = label_font
        ws_summary.cell(row=row, column=2, value=value).font = value_font
        row += 1

    for col in range(1, 5):
        ws_summary.column_dimensions[chr(64 + col)].width = 22

    ws_detail = wb.create_sheet("工单明细")

    detail_headers = ["工单号", "日期时间", "车牌号", "品牌型号", "里程(km)", "故障描述",
                      "故障码", "工时费(元)", "配件费(元)", "总费用(元)"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    from repository import WorkOrderPartRepository
    row = 2
    for order in stats['orders']:
        parts = WorkOrderPartRepository.get_by_work_order_id(order['id'])
        parts_cost = sum(p['subtotal'] for p in parts)
        ws_detail.cell(row=row, column=1, value=order['id']).border = thin_border
        ws_detail.cell(row=row, column=2, value=order['created_at']).border = thin_border
        ws_detail.cell(row=row, column=3, value=order.get('plate_number', '')).border = thin_border
        ws_detail.cell(row=row, column=4, value=order.get('brand_model', '')).border = thin_border
        ws_detail.cell(row=row, column=5, value=order['mileage']).border = thin_border
        ws_detail.cell(row=row, column=6, value=order['fault_description']).border = thin_border
        ws_detail.cell(row=row, column=7, value=order.get('fault_code', '') or '').border = thin_border
        ws_detail.cell(row=row, column=8, value=round(order['labor_cost'], 2)).border = thin_border
        ws_detail.cell(row=row, column=9, value=round(parts_cost, 2)).border = thin_border
        ws_detail.cell(row=row, column=10, value=round(order['total_cost'], 2)).border = thin_border
        row += 1

    col_widths = [10, 20, 12, 18, 12, 30, 12, 12, 12, 12]
    for idx, width in enumerate(col_widths, 1):
        ws_detail.column_dimensions[chr(64 + idx)].width = width

    wb.save(filepath)
    return True, filepath


def generate_history_report(plate_number):
    history = WorkOrderService.get_vehicle_history(plate_number)
    if not history:
        return None

    lines = []
    lines.append("=" * 70)
    lines.append("车辆维修历史报告")
    lines.append("=" * 70)
    lines.append("")

    v = history['vehicle']
    lines.append(f"车牌号:    {v['plate_number']}")
    lines.append(f"品牌型号:  {v['brand_model']}")
    lines.append(f"VIN码:     {v['vin_code']}")
    lines.append(f"上次保养里程: {v['last_maintenance_mileage']} km")
    lines.append(f"建档时间:  {v['created_at']}")
    lines.append("")

    orders = history['orders']
    lines.append(f"维修记录总数: {len(orders)} 次")
    lines.append("-" * 70)

    total_cost_all = 0.0
    for idx, order in enumerate(orders, 1):
        lines.append("")
        lines.append(f"【工单 #{order['id']}】 {order['created_at']}")
        lines.append(f"  进厂里程:   {order['mileage']} km")
        lines.append(f"  故障描述:   {order['fault_description']}")
        if order.get('fault_code'):
            lines.append(f"  故障码:     {order['fault_code']} - {order.get('fault_desc', '')}")
        lines.append("")

        parts = order['parts']
        if parts:
            lines.append("  更换配件:")
            lines.append(f"  {'序号':<6}{'配件名称':<16}{'编号':<10}{'单价(元)':<10}{'数量':<6}{'小计(元)':<10}")
            lines.append("  " + "-" * 58)
            for i, p in enumerate(parts, 1):
                lines.append(f"  {i:<6}{p['part_name']:<16}{p['part_code']:<10}"
                             f"{p['unit_price']:<10.2f}{p['quantity']:<6}{p['subtotal']:<10.2f}")

        parts_subtotal = sum(p['subtotal'] for p in parts)
        lines.append("")
        lines.append(f"  配件费用:   {parts_subtotal:.2f} 元")
        lines.append(f"  工时费用:   {order['labor_cost']:.2f} 元")
        lines.append(f"  工单总计:   {order['total_cost']:.2f} 元")
        total_cost_all += order['total_cost']
        lines.append("")
        lines.append("  " + "-" * 58)

    lines.append("")
    lines.append("=" * 70)
    lines.append(f"累计维修总费用: {total_cost_all:.2f} 元")
    lines.append("=" * 70)

    return "\n".join(lines)


def save_history_report(plate_number, filename=None):
    report = generate_history_report(plate_number)
    if not report:
        return False, "未找到该车辆的维修记录"

    ensure_export_dir()
    if not filename:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"维修历史_{plate_number}_{ts}.txt"
    filepath = os.path.join(EXPORT_DIR, filename)

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(report)

    return True, filepath
